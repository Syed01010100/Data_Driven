# This file is used to READ the excel file (like username,password)

from openpyxl import load_workbook

class Excel_Functions:
    def __init__(self, excel_file, sheet_number):
        self.file = excel_file
        self.sheet = sheet_number

    # fetch the total row count from the excel sheet
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    # fetch the total column count from the excel sheet
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    # read data from excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    # write data to excel file
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)

